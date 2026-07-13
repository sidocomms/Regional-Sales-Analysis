from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


@dataclass(frozen=True)
class WorkbookInfo:
    workbook_path: Path
    sheet_names: list[str]
    selected_sheet: str
    row_count: int
    column_count: int


def load_workbook_info(workbook_path: Path, preferred_sheet: str = "Sales Orders Sheet") -> WorkbookInfo:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    selected_sheet = preferred_sheet if preferred_sheet in sheet_names else sheet_names[0]
    ws = wb[selected_sheet]
    return WorkbookInfo(
        workbook_path=workbook_path,
        sheet_names=sheet_names,
        selected_sheet=selected_sheet,
        row_count=ws.max_row,
        column_count=ws.max_column,
    )


def read_sheet(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(workbook_path, sheet_name=sheet_name)


def load_all_sheets(workbook_path: Path) -> dict[str, pd.DataFrame]:
    xls = pd.ExcelFile(workbook_path)
    return {sheet: pd.read_excel(workbook_path, sheet_name=sheet) for sheet in xls.sheet_names}
